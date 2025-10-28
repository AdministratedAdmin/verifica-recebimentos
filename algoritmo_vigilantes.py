import sys, os
import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle

if getattr(sys, 'frozen', False):
    # Executável
    base_path = sys._MEIPASS
else:
    # Script normal
    base_path = os.path.abspath("")

arquivo_path = os.path.join(base_path, "contracheque_vigilantes.pdf")

wb = Workbook()
ws = wb.active
ws.title = "Planilha pronta"

ws['A1'] = "Nome do vigilante"
ws['B1'] = "Salário" 
ws['C1'] = "Periculosidade paga"
ws['D1'] = "Análise de adicional de periculosidade"

alinhamento_centralizado = Alignment(horizontal='center', vertical='center')

for col in ['A', 'B', 'C', 'D']:
    # Percorre cada célula da coluna e aplica o estilo
    for row in range(1, ws.max_row + 1):
        ws[f'{col}{row}'].alignment = alinhamento_centralizado

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 35

fill_blue = PatternFill(start_color="b3ecff", end_color="b3ecff", fill_type="solid")
ws['A1'].fill = fill_blue
ws['B1'].fill = fill_blue
ws['C1'].fill = fill_blue
ws['D1'].fill = fill_blue

dxf_green = DifferentialStyle(fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
dxf_red = DifferentialStyle(fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))

linhaexcel = 2

with pdfplumber.open(arquivo_path) as pdf:
    # garante que o algoritmo percorra todas as páginas do PDF
    for i, pagina in enumerate(pdf.pages):
        # lê a página atual
        texto = pagina.extract_text()

        # filtro para ignorar outros termos e retornar diretamente o nome
        padraonome = r"""
        \b\d+[.:–\-]?\s+                                # número + separador
        (?!Matr[ií]cula|Data|\nCargo|Fun[cç][aã]o|CPF)    # ignora palavras indesejadas
        ([A-ZÀ-Ú]+(?:\s(?:DA|DE|DO|DOS|DAS|E)?\s?[A-ZÀ-Ú]{2,}){1,3}) # +1 a 3 sobrenomes
        """

        # filtro para retornar a periculosidade paga para fins de visualização
        padraopericulosidade = r"PERICULOSIDADE.*?R\$[\s]?(\d{1,3}(?:\.\d{3})*,\d{2})"

        # filtro para o salário que será comparado na periculosidade
        padraosalario = r"SALARIO MES CIVIL.*?R\$[\s]?(\d{1,3}(?:\.\d{3})*,\d{2})"


        nomes = re.findall(padraonome, texto, re.VERBOSE)
        nomevigilante = nomes[1] if len(nomes) >= 2 else None ## estabelece que o findall deve retornar valores com mais de 2 caracteres

        periculosidadebruta = re.findall(padraopericulosidade, texto)
        periculosidade = [float(v.replace('.', '').replace(',', '.')) for v in periculosidadebruta]

        salariobruto = re.findall(padraosalario, texto)
        salario = [float(s.replace('.', '').replace(',', '.')) for s in salariobruto]

        # Verifica periculosidade
        if salario and periculosidade:
            # Verifica se periculosidade é 30% do salário base
            if round(periculosidade[0],2) == round(salario[0] * 0.3, 2):
                pericuconformidade = "Está em conformidade"
            else:
                pericuconformidade = "Não está em conformidade"
        else:
            pericuconformidade = "Dados insuficientes"

        # Resultados
        print(f"Página {i+1} - Nome: {nomevigilante}")
        print(f"Salário: {salario}")
        print(f"Periculosidade: {periculosidade}")
        print(f"Periculosidade: {pericuconformidade}")
        print("-"*40)

        # Escreve na planilha exatamente o que aparece nos prints
        ws[f'A{linhaexcel}'] = nomevigilante or ""
        ws[f'B{linhaexcel}'] = salario[0] if salario else 0.0
        ws[f'C{linhaexcel}'] = periculosidade[0] if periculosidade else 0.0
        ws[f'D{linhaexcel}'] = pericuconformidade or ""
        linhaexcel += 1  # avança para próxima linha

ruleverdadeiro = Rule(
    type="expression",
    dxf=dxf_green,
    formula=['EXACT(D2, "Está em conformidade")']
)

rulefalso = Rule(
    type="expression",
    dxf=dxf_red,
    formula=['EXACT(D2, "Não está em conformidade")']
)



ws.conditional_formatting.add('D2:D50', ruleverdadeiro)
ws.conditional_formatting.add('D2:D50', rulefalso)

# Salva a planilha
wb.save("conformidade_euroseg.xlsx")
print("Planilha criada com sucesso!")

