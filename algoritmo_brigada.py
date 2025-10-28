import sys, os
import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

if getattr(sys, 'frozen', False):
    # Executável
    base_path = sys._MEIPASS
else:
    # Script normal
    base_path = os.path.abspath(".")

arquivo_path = os.path.join(base_path, "folhadepagamento_brigada.pdf")

todos_nomes = []
todos_salarios = []
todos_periculosidade = []
todas_conformidades = []  # nova lista: conformidade individual

# Padrões regex
padraonome = r'Empr\.\:\s*(\d*[A-ZÀ-Ú][A-ZÀ-Ú\s\-]*)\s*Situação'
padrao_periculosidade = re.compile(
    r'(?i)(?:PERICULOSIDADE)\s*\d{1,3},\d{2}\s+([0-9]{1,3}(?:\.[0-9]{3})*,\d{2})(?=P)'
)

# Função para converter string do formato brasileiro para float
def str_para_float(valor_str):
    if not valor_str:
        return None
    s = valor_str.replace('.', '').replace(',', '.').strip()
    try:
        return float(s)
    except ValueError:
        return None

# Função para extrair periculosidade por página
def extrair_periculosidade_por_pagina(texto, max_por_pagina=5):
    encontrados = padrao_periculosidade.findall(texto)
    resultados = []
    for v in encontrados[:max_por_pagina]:
        f = str_para_float(v)
        if f and f > 0:
            resultados.append(f)
    return resultados

# Loop principal do PDF
with pdfplumber.open(arquivo_path) as pdf:
    for i, pagina in enumerate(pdf.pages):
        texto = pagina.extract_text() or ""

        # Extrai nomes
        nomes_encontrados = re.findall(padraonome, texto)
        nomes_encontrados = [
            re.sub(r'^\d+', '', nome).strip()
            for nome in nomes_encontrados
            if nome.strip()
        ][:5]

        # Ignora página sem nomes
        if not nomes_encontrados:
            continue

        todos_nomes.extend(nomes_encontrados)
        print(f"[Página {i+1}] Nomes:", nomes_encontrados)

        # Extrai salários e converte para float
        salario = re.findall(r'(?i)sal[áa]rio:\s*([\d\.\,]+)', texto)[:5]
        salario_float = [str_para_float(v) for v in salario if str_para_float(v) is not None]
        todos_salarios.extend(salario_float)
        print("Salários:", salario_float)

        # Extrai periculosidade e converte para float
        periculosidade_float = extrair_periculosidade_por_pagina(texto, max_por_pagina=5)
        todos_periculosidade.extend(periculosidade_float)
        print("Periculosidade:", periculosidade_float)

        # Gera conformidade por funcionário
        conformidades_pagina = []
        for idx, nome in enumerate(nomes_encontrados):
            # Pega o salário e periculosidade correspondentes (ou 0 se não houver)
            s = salario_float[idx] if idx < len(salario_float) else 0
            p = periculosidade_float[idx] if idx < len(periculosidade_float) else 0

            if s > 0 and p > 0:
                if round(p, 2) == round(s * 0.3, 2):
                    conformidade = "Está em conformidade"
                else:
                    conformidade = "Não está em conformidade"
            else:
                conformidade = "Dados insuficientes"

            conformidades_pagina.append(conformidade)

        todas_conformidades.extend(conformidades_pagina)
        print("Conformidades:", conformidades_pagina)
        print("-" * 40)

dados = {
    'Nome': todos_nomes,
    'Salário': todos_salarios,
    'Periculosidade': todos_periculosidade,
    'Conformidade': todas_conformidades
}

df = pd.DataFrame(dados)

# Salvar em Excel
arquivo_excel = 'conformidade_brigada.xlsx'
df.to_excel(arquivo_excel, index=False)

# Abre para customizar
wb = load_workbook(arquivo_excel)
ws = wb.active

# 1️⃣ Cabeçalho com background hex b3ecff
fill_cabecalho = PatternFill(start_color="B3ECFF", end_color="B3ECFF", fill_type="solid")
for col in range(1, len(df.columns)+1):
    cell = ws.cell(row=1, column=col)
    cell.fill = fill_cabecalho
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 2️⃣ Largura das colunas: primeira e última 35, demais auto
num_colunas = len(df.columns)
for i, column in enumerate(df.columns, start=1):
    if i == 1 or i == num_colunas:
        ws.column_dimensions[get_column_letter(i)].width = 35
    else:
        max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
        ws.column_dimensions[get_column_letter(i)].width = max_length

# 3️⃣ Formatação condicional na coluna Conformidade (última coluna)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=num_colunas)
    if cell.value == "Está em conformidade":
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde
    elif cell.value == "Não está em conformidade":
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # vermelho

# Salva planilha final
wb.save(arquivo_excel)
print("Planilha customizada criada com sucesso!")